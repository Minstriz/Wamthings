import json
import logging
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# Cấu hình logging
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.DEBUG,
    handlers=[
        logging.FileHandler('excel_processor.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Cấu hình
JSON_FILE = 'attendance_data.json'
EXCEL_FILE = 'attendance_report.xlsx'

def calculate_work_hours(checkin, checkout):
    """Tính số giờ làm việc và giờ tăng ca."""
    try:
        if checkin and checkout:
            checkin_dt = datetime.strptime(checkin, "%Y-%m-%d %H:%M:%S")
            checkout_dt = datetime.strptime(checkout, "%Y-%m-%d %H:%M:%S")
            hours_worked = (checkout_dt - checkin_dt).total_seconds() / 3600.0
            hours_worked = round(hours_worked, 2)
            overtime_hours = max(0, hours_worked - 8)
            overtime_hours = round(overtime_hours, 2)
            work_coeff = 1.0
            overtime_coeff = 1.5 if overtime_hours > 0 else 1.0
            return hours_worked, overtime_hours, work_coeff, overtime_coeff
        return 0, 0, 1.0, 1.0
    except ValueError as e:
        logger.error(f"Invalid time format: {str(e)}")
        return 0, 0, 1.0, 1.0

def generate_excel_report():
    """
    Xử lý dữ liệu JSON, ghép cặp check-in/check-out và tạo file Excel.
    Trả về đường dẫn tới file Excel hoặc None nếu thất bại.
    """
    try:
        # Đọc dữ liệu JSON
        with open(JSON_FILE, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
        
        # Tiêu đề cho Excel
        headers = [
            "Ngày", "Tên", "Check-in", "Check-out",
            "Số giờ làm việc thực tế", "Số giờ tăng ca",
            "Hệ số làm việc thực tế", "Hệ số giờ tăng ca"
        ]
        
        # Ghép cặp check-in và check-out
        data_for_excel = []
        name_to_checkin = {}
        
        for record in json_data:
            name = record.get("Tên", "")
            checkin_time = record.get("Check-in", "")
            checkout_time = record.get("Check-out", "")
            timestamp = record.get("Thời gian", "")
            
            # Lấy ngày từ thời gian
            try:
                date = datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d")
            except ValueError:
                date = ""
            
            if name not in name_to_checkin:
                name_to_checkin[name] = None
            
            if checkin_time and not name_to_checkin[name]:
                name_to_checkin[name] = (checkin_time, date)
                logger.debug(f"Recorded check-in for {name} at {checkin_time}")
            elif checkout_time and name_to_checkin[name]:
                checkin_time, checkin_date = name_to_checkin[name]
                hours_worked, overtime_hours, work_coeff, overtime_coeff = calculate_work_hours(
                    checkin_time, checkout_time
                )
                data_for_excel.append([
                    checkin_date,
                    name,
                    checkin_time,
                    checkout_time,
                    hours_worked,
                    overtime_hours,
                    work_coeff,
                    overtime_coeff
                ])
                logger.debug(f"Paired check-in/check-out for {name}: {checkin_time} -> {checkout_time}")
                name_to_checkin[name] = None
        
        # Tạo DataFrame
        df = pd.DataFrame(data_for_excel, columns=headers)
        
        # Tạo workbook và worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Report"
        
        # Ghi tiêu đề
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )
        
        # Ghi dữ liệu
        for row_idx, row in enumerate(data_for_excel, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal="center")
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin")
                )
        
        # Điều chỉnh độ rộng cột
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Lưu file
        wb.save(EXCEL_FILE)
        logger.info(f"Successfully wrote data to {EXCEL_FILE}")
        return EXCEL_FILE
    
    except Exception as e:
        logger.error(f"Error generating Excel report: {str(e)}")
        return None

if __name__ == "__main__":
    logger.info("Starting Excel report generation test")
    result = generate_excel_report()
    if result:
        logger.info(f"Test successful: Excel file generated at {result}")
    else:
        logger.error("Test failed: Could not generate Excel file")